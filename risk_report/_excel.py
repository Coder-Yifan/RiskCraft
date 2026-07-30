"""独立 Excel 写入器 — 可灵活将任意 DataFrame 或报告章节落地为 Excel。

美化特性:
- 微软雅黑字体
- 交替行背景色
- 百分比列自动格式化
- lift/bad_rate 等列条件格式数据条
- 自动调整列宽（中文字符宽度估算）
"""

from typing import Any

import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.formatting.rule import DataBarRule

from ._base import SubSection
from ._format import FormatConfig, DEFAULT_FORMAT


class ExcelWriter:
    """独立 Excel 写入器。

    与 ModelReport 无耦合，支持三种写入方式:
    1. write_sub_sections() — 写入 list[SubSection] 到指定 sheet
    2. write_dataframe() — 写入单个 DataFrame（可选标题）
    3. write_report() — 写入完整报告（dict[sheet_name, list[SubSection]] → 多个 sheet）

    Parameters
    ----------
    file_path : str
        输出文件路径
    format_config : FormatConfig | None
        全局格式化配置，默认 DEFAULT_FORMAT
    auto_adjust_width : bool
        是否自动调整列宽，默认 True
    """

    def __init__(
        self,
        file_path: str,
        format_config: FormatConfig | None = None,
        auto_adjust_width: bool = True,
    ):
        self.file_path = file_path
        self.format_config = format_config or DEFAULT_FORMAT
        self.auto_adjust_width = auto_adjust_width
        self._wb = openpyxl.Workbook()
        # 删除默认 sheet
        if "Sheet" in self._wb.sheetnames:
            self._wb.remove(self._wb["Sheet"])

    def write_sub_sections(
        self,
        sub_sections: list[SubSection],
        sheet_name: str,
        format_config: FormatConfig | None = None,
    ) -> None:
        """写入 list[SubSection] 到指定 sheet。"""
        # 确保唯一 sheet 名
        name = sheet_name
        if name in self._wb.sheetnames:
            name = f"{name}_{len(self._wb.sheetnames)}"
        ws = self._wb.create_sheet(name)

        fmt = format_config or self.format_config
        row = 1

        for sub in sub_sections:
            row = self._write_sub_section(ws, sub, row, fmt, sheet_name=name)
            row += fmt.subsection_gap_rows

    def write_dataframe(
        self,
        df: pd.DataFrame,
        sheet_name: str,
        start_row: int = 1,
        format_config: FormatConfig | None = None,
        title: str | None = None,
    ) -> None:
        """写入单个 DataFrame 到指定 sheet。"""
        """写入单个 DataFrame 到指定 sheet。"""
        if sheet_name not in self._wb.sheetnames:
            ws = self._wb.create_sheet(sheet_name)
        else:
            ws = self._wb[sheet_name]

        fmt = format_config or self.format_config
        row = start_row

        if title:
            ws.cell(row=row, column=1, value=title).font = Font(
                name=fmt.font_name, size=fmt.title_font_size,
                bold=fmt.title_font_bold,
            )
            row += 1

        row = self._write_dataframe_rows(ws, df, row, fmt, sheet_name=sheet_name)

        if self.auto_adjust_width:
            self._adjust_column_width(ws)

    def write_report(self, sheet_results: dict[str, list[SubSection]]) -> None:
        """写入完整报告（每个 sheet 一个 SubSection 列表）。"""
        for sheet_name, sub_sections in sheet_results.items():
            self.write_sub_sections(sub_sections, sheet_name)

    def save(self) -> None:
        """保存到 file_path。"""
        self._wb.save(self.file_path)

    # ---- 内部方法 ----

    def _write_sub_section(
        self,
        ws: Worksheet,
        sub: SubSection,
        start_row: int,
        fmt: FormatConfig,
        sheet_name: str = "",
    ) -> int:
        """写入单个 SubSection，返回结束行号。"""
        row = start_row

        # 标题行
        title_cell = ws.cell(row=row, column=1, value=sub.title)
        title_cell.font = Font(
            name=fmt.font_name, size=fmt.title_font_size,
            bold=fmt.title_font_bold,
        )
        row += 1

        # 备注（如有）
        if sub.note:
            ws.cell(row=row, column=1, value=sub.note).font = Font(
                name=fmt.font_name, size=9, italic=True, color="808080",
            )
            row += 1

        # 数据表
        row = self._write_dataframe_rows(ws, sub.data, row, fmt, sheet_name=sheet_name)

        # 条件格式数据条
        self._add_data_bars(ws, sub.data, start_row + (1 + (1 if sub.note else 0)), fmt)

        if self.auto_adjust_width:
            self._adjust_column_width(ws)

        return row

    def _write_dataframe_rows(
        self,
        ws: Worksheet,
        df: pd.DataFrame,
        start_row: int,
        fmt: FormatConfig,
        sheet_name: str = "",
    ) -> int:
        """写入 DataFrame 表头+数据行，带格式化。"""
        row = start_row
        n_cols = len(df.columns)
        header_row = row

        # ---- 边框 ----
        thin_border = Border(
            left=Side(style=fmt.border_style),
            right=Side(style=fmt.border_style),
            top=Side(style=fmt.border_style),
            bottom=Side(style=fmt.border_style),
        )

        # ---- 表头 ----
        header_fill = PatternFill(
            start_color=fmt.header_bg_color,
            end_color=fmt.header_bg_color,
            fill_type="solid",
        )
        header_font = Font(
            name=fmt.font_name, color=fmt.header_font_color,
            bold=True, size=fmt.data_font_size,
        )
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=row, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
        row += 1

        # ---- 交替行色 ----
        alt_fill = PatternFill(
            start_color=fmt.alt_row_color,
            end_color=fmt.alt_row_color,
            fill_type="solid",
        ) if fmt.alt_row_enabled else None

        # ---- 数据行 ----
        data_font = Font(name=fmt.font_name, size=fmt.data_font_size)
        data_align = Alignment(horizontal="center", vertical="center")

        # 预计算百分比列
        percent_cols = set()
        for col_idx, col_name in enumerate(df.columns, 1):
            if self._is_percent_column(col_name, fmt):
                percent_cols.add(col_idx)

        for row_idx, (_, data_row) in enumerate(df.iterrows()):
            is_alt = (row_idx % 2 == 1)
            for col_idx, val in enumerate(data_row, 1):
                # NaN → 空字符串
                if val is None or (isinstance(val, float) and np.isnan(val)):
                    cell_value = ""
                else:
                    cell_value = val

                cell = ws.cell(row=row, column=col_idx, value=cell_value)
                cell.font = data_font
                cell.alignment = data_align
                cell.border = thin_border

                # 交替行色
                if is_alt and alt_fill is not None:
                    cell.fill = alt_fill

                # 数字格式
                if isinstance(cell_value, (int, np.integer)):
                    cell.number_format = fmt.integer_format
                elif isinstance(cell_value, float) and not np.isnan(cell_value if isinstance(val, float) else 0):
                    if col_idx in percent_cols:
                        cell.number_format = fmt.percent_format
                    else:
                        cell.number_format = fmt.float_format
            row += 1

        # 冻结表头（仅指定 Sheet）
        if any(s in sheet_name for s in fmt.freeze_header_sheets):
            ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

        return row

    def _add_data_bars(
        self,
        ws: Worksheet,
        df: pd.DataFrame,
        header_row: int,
        fmt: FormatConfig,
    ) -> None:
        """为指定列添加条件格式数据条。"""
        if not fmt.data_bar_columns:
            return

        data_start = header_row + 1
        data_end = data_start + len(df) - 1
        if data_end < data_start:
            return

        for col_idx, col_name in enumerate(df.columns, 1):
            # 检查列名是否匹配数据条关键字
            col_lower = col_name.lower()
            if not any(kw in col_lower for kw in fmt.data_bar_columns):
                continue

            # 检查列是否为数值型
            if not pd.api.types.is_numeric_dtype(df[col_name]):
                continue

            col_letter = get_column_letter(col_idx)
            cell_range = f"{col_letter}{data_start}:{col_letter}{data_end}"

            rule = DataBarRule(
                start_type="min", start_value=0,
                end_type="max", end_value=0,
                color="6BAED6",  # 柔和蓝色
                showValue=True,
            )
            ws.conditional_formatting.add(cell_range, rule)

    def _is_percent_column(self, col_name: str, fmt: FormatConfig) -> bool:
        """判断列是否为百分比列。"""
        col_lower = col_name.lower()
        # 1. 列名以 % 结尾
        if col_name.endswith("%"):
            return True
        # 2. 列名匹配 percent_columns 关键字
        if any(kw.lower() in col_lower for kw in fmt.percent_columns):
            return True
        return False

    def _adjust_column_width(self, ws: Worksheet) -> None:
        """根据内容自动调整列宽。"""
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        val_str = str(cell.value)
                        length = len(val_str)
                        # 中文字符占2个宽度单位
                        cn_chars = sum(1 for c in val_str if '一' <= c <= '鿿')
                        length = length + cn_chars
                        max_length = max(max_length, length)
                except Exception:
                    pass
            adjusted_width = min(max(max_length + 2, 8), 50)
            ws.column_dimensions[col_letter].width = adjusted_width
