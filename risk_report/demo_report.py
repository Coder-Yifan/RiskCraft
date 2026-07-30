"""risk_report 模块完整演示 — 使用 demo_data.csv 数据集。

演示流程:
1. 加载 demo_data.csv
2. 按 transaction_time 划分 train/test/oot，写入 tag 列
3. 构建风控建模 Pipeline 并训练
4. 构造 ReportContext（单 DataFrame + tag_col + label_col）
5. 产出完整模型开发文档 Excel（8 Sheet / 22 算子）
"""

import sys
import time
import numpy as np
import pandas as pd

sys.stdout.reconfigure(encoding="utf-8")

print("=" * 60)
print("RiskCraft — risk_report 模块完整演示")
print("=" * 60)

# ============================================================
# 1. 加载数据
# ============================================================
print("\n[1] 加载 demo_data.csv ...")
df = pd.read_csv("risk_ml/dataset/demo_data.csv")
print(f"  总样本: {len(df)}, 欺诈率: {df['is_fraud'].mean():.4f}")

# 解析时间列
df["transaction_time"] = pd.to_datetime(df["transaction_time"])
print(f"  时间范围: {df['transaction_time'].min()} ~ {df['transaction_time'].max()}")

# ============================================================
# 2. 划分数据集 — 写入 tag 列
# ============================================================
print("\n[2] 按时间划分 train / test / OOT → 写入 tag 列 ...")

# 特征列: 排除 ID/时间/标签 和高基数类别列
exclude_cols = [
    "transaction_id", "customer_id", "transaction_time",
    "is_fraud",  # 标签
    "monthly_spend",  # 高基数, 近似唯一值
]
feature_cols = [c for c in df.columns if c not in exclude_cols]
print(f"  原始特征数: {len(df.columns) - len(exclude_cols)} → {len(feature_cols)}")

# 时间划分 → 写入 tag 列
train_mask = df["transaction_time"] < "2024-01-01"
test_mask = (df["transaction_time"] >= "2024-01-01") & (df["transaction_time"] < "2024-07-01")
oot_mask = df["transaction_time"] >= "2024-07-01"

df["tag"] = "train"
df.loc[test_mask, "tag"] = "test"
df.loc[oot_mask, "tag"] = "oot"

print(f"  train: {(df['tag']=='train').sum()} 条, 欺诈率 {df.loc[df['tag']=='train', 'is_fraud'].mean():.4f}")
print(f"  test:  {(df['tag']=='test').sum()} 条, 欺诈率 {df.loc[df['tag']=='test', 'is_fraud'].mean():.4f}")
print(f"  oot:   {(df['tag']=='oot').sum()} 条, 欺诈率 {df.loc[df['tag']=='oot', 'is_fraud'].mean():.4f}")

# ============================================================
# 3. 构建 Pipeline 并训练
# ============================================================
print("\n[3] 构建风控建模 Pipeline 并训练 ...")

from sklearn.pipeline import Pipeline
from risk_ml.preprocessing import FeatureCleaner
from risk_ml.encoding import BinnerWoeEncoder
from risk_ml.feature_selection import IVSelector, CorrelationSelector
from risk_ml.estimator import RiskXGBClassifier, OptunaTuner

pipe = Pipeline([
    ("cleaner", FeatureCleaner()),
    ("binner_woe", BinnerWoeEncoder(max_bins=8)),
    ("iv_selector", IVSelector(iv_threshold=0.02)),
    ("corr_selector", CorrelationSelector(corr_threshold=0.7)),
    ("classifier", OptunaTuner(
        estimator=RiskXGBClassifier(),
        n_trials=20,
        scoring="ks",
        cv=3,
        random_state=42,
        verbose=0,
    )),
])

start = time.time()
X_train = df.loc[df["tag"] == "train", feature_cols]
y_train = df.loc[df["tag"] == "train", "is_fraud"].values
pipe.fit(X_train, y_train)
elapsed = time.time() - start
print(f"  训练完成, 耗时 {elapsed:.1f}s")

tuner = pipe.named_steps["classifier"]
print(f"  Optuna 最优参数: {tuner.best_params_}")
print(f"  Optuna 最优 KS: {tuner.best_score_:.4f}")

# ============================================================
# 4. 构造 ReportContext（单 DataFrame + tag_col + label_col）
# ============================================================
print("\n[4] 构造 ReportContext（单 DataFrame + tag_col + label_col） ...")

from risk_report import ReportContext

feature_meta = {
    "transaction_amount": {"含义": "交易金额", "来源": "交易系统", "类别": "数值"},
    "merchant_category": {"含义": "商户类别", "来源": "商户信息", "类别": "分类"},
    "transaction_type": {"含义": "交易类型", "来源": "交易系统", "类别": "分类"},
    "payment_method": {"含义": "支付方式", "来源": "交易系统", "类别": "分类"},
    "city": {"含义": "城市", "来源": "用户信息", "类别": "分类"},
    "country": {"含义": "国家", "来源": "用户信息", "类别": "分类"},
    "device_type": {"含义": "设备类型", "来源": "交易日志", "类别": "分类"},
    "operating_system": {"含义": "操作系统", "来源": "交易日志", "类别": "分类"},
    "browser": {"含义": "浏览器", "来源": "交易日志", "类别": "分类"},
    "card_type": {"含义": "卡类型", "来源": "卡片信息", "类别": "分类"},
    "card_present": {"含义": "是否现场刷卡", "来源": "交易系统", "类别": "二元"},
    "international_transaction": {"含义": "是否跨境交易", "来源": "交易系统", "类别": "二元"},
    "distance_from_home": {"含义": "离家距离", "来源": "位置数据", "类别": "数值"},
    "previous_transaction_gap": {"含义": "上次交易间隔", "来源": "交易日志", "类别": "数值"},
    "daily_transaction_count": {"含义": "当日交易次数", "来源": "交易日志", "类别": "数值"},
    "risk_score": {"含义": "风控评分", "来源": "风控引擎", "类别": "数值"},
    "customer_age": {"含义": "客户年龄", "来源": "客户信息", "类别": "数值"},
    "account_tenure_years": {"含义": "账户年限", "来源": "客户信息", "类别": "数值"},
    "merchant_risk_level": {"含义": "商户风险等级", "来源": "风控引擎", "类别": "分类"},
}

context = ReportContext(
    data=df,
    tag_col="tag",
    label_col="is_fraud",
    pipeline=pipe,
    model_name="反欺诈模型_v1.0",
    developer="RiskCraft Demo",
    validator="风控团队",
    business_owner="业务运营部",
    background="针对线上交易欺诈风险，构建反欺诈预测模型，降低欺诈损失率",
    application="线上交易实时风控筛查，辅助人工审核决策",
    label_definition={0: "正常", 1: "欺诈"},
    feature_meta=feature_meta,
)

print(f"  pipeline_attrs.feature_names_in_: {context.pipeline_attrs.feature_names_in_}")
print(f"  score_col: {context.score_col}")
print(f"  get_datasets(): {list(context.get_datasets().keys())}")

# ============================================================
# 5. 产出完整模型开发文档（8 Sheet / 22 算子）
# ============================================================
print("\n[5] 产出完整模型开发文档 Excel ...")

from risk_report import ModelReport

report = ModelReport()
report.fit(context)

print(f"  生成算子结果: {list(report.results_.keys())}")

output_path = "risk_report/demo_report_output.xlsx"
report.to_excel(output_path)
print(f"  ✓ Excel 已保存: {output_path}")

# 验证 Excel 内容
import openpyxl
wb = openpyxl.load_workbook(output_path)
print(f"\n  Excel Sheet 概览:")
print(f"  {'Sheet名':<20s} {'行数':>6s} {'列数':>6s}")
print(f"  {'-'*20} {'-'*6} {'-'*6}")
for name in wb.sheetnames:
    ws = wb[name]
    print(f"  {name:<20s} {ws.max_row:>6d} {ws.max_column:>6d}")

# ============================================================
# 6. 快速查看各数据集指标
# ============================================================
print("\n[6] 各数据集模型效果:")
from risk_report import ModelEffectOperator

datasets = context.get_datasets()
effect_df = ModelEffectOperator.compute_effect_table(datasets)
print(effect_df.to_string(index=False))

# ============================================================
# 7. 自定义配置示例
# ============================================================
print("\n[7] 自定义配置示例 — 只产出模型表现 sheet ...")
from risk_report import SheetConfig, DocumentConfig, ScoreLiftOperator

custom_config = DocumentConfig(sheets=[
    SheetConfig("3.模型表现", [ModelMethodOperator(), ModelEffectOperator(), ScoreLiftOperator()]),
])
custom_report = ModelReport(config=custom_config)
custom_report.fit(context)

output_path2 = "risk_report/demo_report_custom.xlsx"
custom_report.to_excel(output_path2)
print(f"  ✓ 自定义 Excel 已保存: {output_path2}")

print("\n" + "=" * 60)
print("演示完成! 报告文件:")
print("  - risk_report/demo_report_output.xlsx (全量)")
print("  - risk_report/demo_report_custom.xlsx (自定义)")
print("=" * 60)
