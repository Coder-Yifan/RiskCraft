"""risk_report 模块单元测试 — 新架构（单 DataFrame + tag 列 + 22 算子）。"""

import numpy as np
import pandas as pd
import pytest
import tempfile
import os

from risk_report import (
    ReportContext, PipelineAttributes, ModelReport, ExcelWriter,
    ScoreLiftOperator, ModelEffectOperator, SwapAnalysisOperator,
    compute_lift_table, compute_swap_analysis, compute_per_feature_ks, compute_sample_stats,
    MetaInfoOperator, DevPurposeOperator, ModelAssumptionOperator,
    LabelDefinitionOperator, SampleSelectionOperator, ModelingSampleOperator,
    EffectSummaryOperator, VarDescriptionOperator, VarCleaningOperator,
    VarFilterOperator, VarAnalysisOperator, VarBinningOperator,
    ModelMethodOperator, ScoreLiftGrayOperator,
    AttributionOperator, ModelComparisonOperator, MobPerformanceOperator,
    PortraitOperator, VarRangeOperator,
    SubSection, SheetConfig, DocumentConfig,
    FormatConfig, DEFAULT_FORMAT, DEFAULT_DOCUMENT_CONFIG,
    extract_pipeline_attributes, placeholder_df, TAG_CN_MAP,
)


# ============================================================
# 辅助函数: 构造测试数据
# ============================================================

def _make_test_data(n=300, with_gray=False, with_baseline=False, with_extra_labels=False):
    """构造测试 DataFrame（含 tag 列、标签列、特征列）。"""
    np.random.seed(42)
    n_train = int(n * 0.6)
    n_test = int(n * 0.2)
    n_oot = n - n_train - n_test

    features = {
        "f1": np.random.randn(n),
        "f2": np.random.randn(n),
        "f3": np.random.randn(n),
        "f4": np.random.randn(n),
        "f5": np.random.randn(n),
    }
    df = pd.DataFrame(features)
    y = (df["f1"] + df["f2"] > 0).astype(int)
    df["is_default"] = y.values
    df["tag"] = ["train"] * n_train + ["test"] * n_test + ["oot"] * n_oot

    if with_baseline:
        df["baseline_score"] = np.random.rand(n) * 0.5 + 0.2

    if with_gray:
        n_gray = 50
        gray_df = pd.DataFrame({
            "f1": np.random.randn(n_gray),
            "f2": np.random.randn(n_gray),
            "f3": np.random.randn(n_gray),
            "f4": np.random.randn(n_gray),
            "f5": np.random.randn(n_gray),
            "is_default": [-1] * n_gray,
            "tag": ["gray"] * n_gray,
        })
        if with_baseline:
            gray_df["baseline_score"] = np.random.rand(n_gray) * 0.3 + 0.1
        df = pd.concat([df, gray_df], ignore_index=True)

    if with_extra_labels:
        df["y_mob1"] = ((df["f1"] > 0) & (df["is_default"] == 1)).astype(int)
        df["y_mob2"] = ((df["f2"] > 0) & (df["is_default"] == 1)).astype(int)

    return df


def _make_test_pipeline(df):
    """构造并拟合测试 pipeline，返回 (pipeline, fitted_df_with_score)。"""
    from sklearn.pipeline import Pipeline
    from risk_ml.preprocessing import FeatureCleaner
    from risk_ml.encoding import BinnerWoeEncoder
    from risk_ml.feature_selection import IVSelector
    from risk_ml.estimator import RiskXGBClassifier

    mask = df["tag"] == "train"
    X = df[["f1", "f2", "f3", "f4", "f5"]]
    y = df.loc[mask, "is_default"].values.astype(float)

    pipe = Pipeline([
        ("cleaner", FeatureCleaner()),
        ("binner_woe", BinnerWoeEncoder()),
        ("iv_selector", IVSelector()),
        ("classifier", RiskXGBClassifier(n_estimators=20)),
    ])
    pipe.fit(X[mask], y)
    return pipe


def _make_context(with_gray=False, with_baseline=False, with_extra_labels=False, n=300):
    """构造完整 ReportContext（含 pipeline + data）。"""
    df = _make_test_data(n, with_gray=with_gray, with_baseline=with_baseline, with_extra_labels=with_extra_labels)
    pipe = _make_test_pipeline(df)

    ctx = ReportContext(
        data=df,
        tag_col="tag",
        label_col="is_default",
        pipeline=pipe,
        model_name="测试模型",
        developer="dev",
        background="降低违约率",
        application="信贷审批",
        gray_tag="gray" if with_gray else None,
        baseline_score_col="baseline_score" if with_baseline else None,
        extra_labels=["y_mob1", "y_mob2"] if with_extra_labels else None,
    )
    return ctx


# ============================================================
# _scoring.py 测试（保持不变）
# ============================================================

class TestComputeLiftTable:
    def test_basic_output(self):
        y_true = np.array([0, 0, 0, 1, 1, 1])
        y_score = np.array([0.1, 0.3, 0.5, 0.7, 0.8, 0.9])
        df = compute_lift_table(y_true, y_score, n_bins=3)
        assert isinstance(df, pd.DataFrame)
        expected_cols = ["min", "max", "goods", "bads", "total", "total%", "bad_rate", "ks", "lift", "cum_lift"]
        for col in expected_cols:
            assert col in df.columns

    def test_with_baseline(self):
        y_true = np.array([0, 0, 1, 1])
        y_score = np.array([0.2, 0.4, 0.6, 0.8])
        baseline = np.array([0.3, 0.5, 0.5, 0.7])
        df = compute_lift_table(y_true, y_score, n_bins=2, baseline_score=baseline)
        assert "baseline_goods" in df.columns
        assert "baseline_bads" in df.columns


class TestComputeSwapAnalysis:
    def test_basic_no_baseline(self):
        y_true = np.array([0, 0, 1, 1])
        y_score = np.array([0.2, 0.5, 0.7, 0.9])
        df = compute_swap_analysis(y_true, y_score, cutoff_percentiles=[10])
        assert "切分比例" in df.columns
        assert "新总拒绝" in df.columns

    def test_with_baseline(self):
        y_true = np.array([0, 0, 1, 1])
        y_score_new = np.array([0.2, 0.5, 0.7, 0.9])
        y_score_old = np.array([0.3, 0.4, 0.6, 0.8])
        df = compute_swap_analysis(y_true, y_score_new, y_score_old, cutoff_percentiles=[10, 20])
        assert "swap_in" in df.columns
        assert len(df) == 2


class TestComputeSampleStats:
    def test_binary(self):
        y = np.array([0, 0, 1, 1, 1])
        stats = compute_sample_stats(y)
        assert stats["goods"] == 2
        assert stats["bads"] == 3
        assert stats["total"] == 5
        assert stats["bad_rate"] == 0.6

    def test_with_gray(self):
        y = np.array([0, 0, -1, 1, 1])
        stats = compute_sample_stats(y, {0: "好", -1: "灰", 1: "坏"})
        assert stats["gray"] == 1


# ============================================================
# placeholder_df 测试
# ============================================================

class TestPlaceholderDF:
    def test_returns_dataframe(self):
        df = placeholder_df("test message")
        assert isinstance(df, pd.DataFrame)
        assert "说明" in df.columns
        assert df.iloc[0, 0] == "test message"


# ============================================================
# ReportContext 测试（新架构）
# ============================================================

class TestReportContext:
    def test_empty_context(self):
        ctx = ReportContext()
        assert ctx.pipeline_attrs is None
        assert ctx.metrics is not None
        assert ctx.data is None

    def test_get_datasets(self):
        """get_datasets() 从 tag 列拆分数据。"""
        ctx = _make_context()
        datasets = ctx.get_datasets()
        assert "训练集" in datasets
        assert "测试集" in datasets
        assert "跨时间验证集" in datasets
        for cn_name, (y_true, y_score) in datasets.items():
            assert len(y_true) > 0
            assert len(y_score) > 0

    def test_get_baseline_datasets(self):
        """get_baseline_datasets() 从 baseline_score_col 提取。"""
        ctx = _make_context(with_baseline=True)
        datasets = ctx.get_baseline_datasets()
        assert len(datasets) > 0

    def test_get_baseline_datasets_none(self):
        """无 baseline 时返回空字典。"""
        ctx = _make_context()
        datasets = ctx.get_baseline_datasets()
        assert datasets == {}

    def test_get_gray_datasets(self):
        """get_gray_datasets() 从 gray_tag 提取灰样本。"""
        ctx = _make_context(with_gray=True)
        gray = ctx.get_gray_datasets()
        assert "灰样本" in gray

    def test_get_datasets_with_gray(self):
        """get_datasets_with_gray() 合并灰样本。"""
        ctx = _make_context(with_gray=True)
        combined = ctx.get_datasets_with_gray()
        assert any("含灰" in name for name in combined.keys())

    def test_get_sample_stats(self):
        """get_sample_stats() 统计各数据集好坏占比。"""
        ctx = _make_context()
        stats = ctx.get_sample_stats()
        assert "训练集" in stats
        assert stats["训练集"]["goods"] > 0

    def test_auto_score_computation(self):
        """pipeline 自动计算分数写入 data[score_col]。"""
        ctx = _make_context()
        assert ctx.score_col is not None
        assert ctx.score_col in ctx.data.columns


# ============================================================
# ExcelWriter 测试（适配 SheetConfig）
# ============================================================

class TestExcelWriter:
    def test_write_dataframe(self):
        df = pd.DataFrame({"col_a": [1, 2, 3], "col_b": [0.1, 0.2, 0.3]})
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            writer = ExcelWriter(path)
            writer.write_dataframe(df, "test_sheet", title="测试标题")
            writer.save()
            import openpyxl
            wb = openpyxl.load_workbook(path)
            assert "test_sheet" in wb.sheetnames
        finally:
            os.unlink(path)

    def test_write_sub_sections(self):
        subs = [
            SubSection(title="子章节1", data=pd.DataFrame({"x": [1, 2]})),
            SubSection(title="子章节2", data=pd.DataFrame({"y": [3, 4]})),
        ]
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            writer = ExcelWriter(path)
            writer.write_sub_sections(subs, "测试")
            writer.save()
            import openpyxl
            wb = openpyxl.load_workbook(path)
            assert "测试" in wb.sheetnames
        finally:
            os.unlink(path)

    def test_write_report(self):
        sheet_results = {
            "Sheet1": [SubSection("标题", pd.DataFrame({"a": [1]}))],
            "Sheet2": [SubSection("标题2", pd.DataFrame({"b": [2]}))],
        }
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            writer = ExcelWriter(path)
            writer.write_report(sheet_results)
            writer.save()
            import openpyxl
            wb = openpyxl.load_workbook(path)
            assert "Sheet1" in wb.sheetnames
            assert "Sheet2" in wb.sheetnames
        finally:
            os.unlink(path)


# ============================================================
# 算子测试
# ============================================================

class TestMetaInfoOperator:
    def test_compute(self):
        ctx = ReportContext(model_name="测试", developer="dev")
        op = MetaInfoOperator()
        result = op.compute(ctx)
        assert isinstance(result, list)
        assert len(result) == 1
        assert "项目" in result[0].data.columns


class TestDevPurposeOperator:
    def test_compute(self):
        ctx = ReportContext(background="降低违约率")
        result = DevPurposeOperator().compute(ctx)
        assert len(result) == 1
        assert result[0].title == "1.1模型开发目的"


class TestModelAssumptionOperator:
    def test_compute_no_pipeline(self):
        ctx = ReportContext()
        result = ModelAssumptionOperator().compute(ctx)
        assert len(result) == 1
        # 应为占位表
        assert "说明" in result[0].data.columns

    def test_compute_with_pipeline(self):
        ctx = _make_context()
        result = ModelAssumptionOperator().compute(ctx)
        assert len(result) == 1
        assert "参数名称" in result[0].data.columns


class TestLabelDefinitionOperator:
    def test_compute(self):
        ctx = ReportContext(label_definition={0: "好", -1: "灰", 1: "坏"})
        result = LabelDefinitionOperator().compute(ctx)
        assert len(result) >= 1
        assert "标签" in result[0].data.columns


class TestModelingSampleOperator:
    def test_compute(self):
        ctx = _make_context()
        result = ModelingSampleOperator().compute(ctx)
        assert len(result) >= 1
        assert "样本集" in result[0].data.columns
        assert "训练集" in result[0].data["样本集"].values

    def test_compute_with_extra_labels(self):
        ctx = _make_context(with_extra_labels=True)
        result = ModelingSampleOperator().compute(ctx)
        assert len(result) >= 2  # 主标签 + 压测集


class TestScoreLiftOperator:
    def test_static_method(self):
        y_true = np.array([0, 0, 1, 1])
        y_score = np.array([0.2, 0.4, 0.6, 0.8])
        df = ScoreLiftOperator.compute_lift_table(y_true, y_score, n_bins=2)
        assert isinstance(df, pd.DataFrame)
        assert "min" in df.columns

    def test_compute_via_context(self):
        ctx = _make_context()
        op = ScoreLiftOperator(n_bins=5)
        result = op.compute(ctx)
        assert isinstance(result, list)
        assert len(result) > 0
        # 每个 SubSection 对应一个数据集
        assert result[0].title in ["训练集", "测试集", "跨时间验证集"]

    def test_compute_no_data(self):
        ctx = ReportContext()
        result = ScoreLiftOperator().compute(ctx)
        assert len(result) == 1
        assert "说明" in result[0].data.columns  # 占位表


class TestScoreLiftGrayOperator:
    def test_compute_with_gray(self):
        ctx = _make_context(with_gray=True)
        result = ScoreLiftGrayOperator().compute(ctx)
        assert isinstance(result, list)
        assert len(result) > 0

    def test_compute_no_gray(self):
        ctx = _make_context()
        result = ScoreLiftGrayOperator().compute(ctx)
        assert len(result) == 1
        assert "说明" in result[0].data.columns


class TestModelEffectOperator:
    def test_static_method(self):
        datasets = {"train": (np.array([0, 0, 1, 1]), np.array([0.2, 0.4, 0.6, 0.8]))}
        df = ModelEffectOperator.compute_effect_table(datasets)
        assert "数据集" in df.columns
        assert "auc" in df.columns

    def test_compute_via_context(self):
        ctx = _make_context()
        result = ModelEffectOperator().compute(ctx)
        assert len(result) == 1
        assert "数据集" in result[0].data.columns


class TestSwapAnalysisOperator:
    def test_static_method(self):
        y_true = np.array([0, 0, 1, 1])
        y_score = np.array([0.2, 0.5, 0.7, 0.9])
        df = SwapAnalysisOperator.compute_swap_table(y_true, y_score)
        assert "切分比例" in df.columns

    def test_compute_via_context(self):
        ctx = _make_context()
        result = SwapAnalysisOperator().compute(ctx)
        assert isinstance(result, list)
        assert len(result) > 0


class TestVarFilterOperator:
    def test_compute(self):
        ctx = _make_context()
        result = VarFilterOperator().compute(ctx)
        assert len(result) == 1
        assert "筛选指标" in result[0].data.columns


class TestVarRangeOperator:
    def test_compute(self):
        ctx = _make_context()
        result = VarRangeOperator().compute(ctx)
        assert len(result) == 1
        assert "变量名" in result[0].data.columns


class TestVarAnalysisOperator:
    def test_compute(self):
        ctx = _make_context()
        result = VarAnalysisOperator().compute(ctx)
        assert len(result) == 1
        assert "feature" in result[0].data.columns


class TestAttributionOperator:
    def test_compute(self):
        ctx = ReportContext()
        result = AttributionOperator().compute(ctx)
        assert len(result) == 1
        assert "维度" in result[0].data.columns


class TestModelMethodOperator:
    def test_compute_with_pipeline(self):
        ctx = _make_context()
        result = ModelMethodOperator().compute(ctx)
        assert len(result) == 1
        assert "参数名称" in result[0].data.columns


# ============================================================
# 配置测试
# ============================================================

class TestSheetConfig:
    def test_creation(self):
        cfg = SheetConfig(sheet_name="测试", operators=[MetaInfoOperator()])
        assert cfg.sheet_name == "测试"
        assert len(cfg.operators) == 1


class TestDocumentConfig:
    def test_default_config(self):
        cfg = DEFAULT_DOCUMENT_CONFIG
        assert len(cfg.sheets) == 8  # 8 个 Sheet
        # 22 个算子
        total_ops = sum(len(s.operators) for s in cfg.sheets)
        assert total_ops == 22


class TestCustomConfig:
    def test_custom_document(self):
        """自定义 DocumentConfig 选择部分算子。"""
        cfg = DocumentConfig(sheets=[
            SheetConfig("测试Sheet", [MetaInfoOperator(), DevPurposeOperator()]),
        ])
        ctx = ReportContext(model_name="自定义测试")
        report = ModelReport(config=cfg)
        report.fit(ctx)
        assert "meta_info" in report.results_
        assert "dev_purpose" in report.results_


# ============================================================
# ModelReport 组合器测试
# ============================================================

class TestModelReport:
    def test_default_config(self):
        report = ModelReport()
        assert len(report.config.sheets) == 8

    def test_fit_and_to_excel(self):
        """完整流程: fit → to_excel。"""
        ctx = _make_context()
        # 只用部分算子加速测试
        cfg = DocumentConfig(sheets=[
            SheetConfig("模型说明", [MetaInfoOperator()]),
            SheetConfig("1.模型设计", [DevPurposeOperator(), ModelingSampleOperator()]),
        ])
        report = ModelReport(config=cfg)
        report.fit(ctx)

        assert "meta_info" in report.results_
        assert "dev_purpose" in report.results_

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            report.to_excel(path)
            import openpyxl
            wb = openpyxl.load_workbook(path)
            assert len(wb.sheetnames) == 2
        finally:
            os.unlink(path)

    def test_get_result(self):
        ctx = ReportContext(model_name="test")
        cfg = DocumentConfig(sheets=[SheetConfig("测试", [MetaInfoOperator()])])
        report = ModelReport(config=cfg)
        report.fit(ctx)
        result = report.get_result("meta_info")
        assert isinstance(result, list)
        assert len(result) == 1

    def test_get_result_not_found(self):
        ctx = ReportContext(model_name="test")
        cfg = DocumentConfig(sheets=[SheetConfig("测试", [MetaInfoOperator()])])
        report = ModelReport(config=cfg)
        report.fit(ctx)
        with pytest.raises(KeyError):
            report.get_result("nonexistent")

    def test_full_report_all_operators(self):
        """全量报告: 22 个算子 fit + to_excel。"""
        ctx = _make_context(with_baseline=True, with_gray=True, with_extra_labels=True)
        report = ModelReport()
        report.fit(ctx)

        # 22 个算子都应运行
        assert len(report.results_) == 22

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            report.to_excel(path)
            import openpyxl
            wb = openpyxl.load_workbook(path)
            # 8 个 sheet
            assert len(wb.sheetnames) == 8
        finally:
            os.unlink(path)


# ============================================================
# TAG_CN_MAP 测试
# ============================================================

class TestTagCNMap:
    def test_mapping(self):
        assert TAG_CN_MAP["train"] == "训练集"
        assert TAG_CN_MAP["test"] == "测试集"
        assert TAG_CN_MAP["oot"] == "跨时间验证集"


# ============================================================
# extract_pipeline_attributes 测试
# ============================================================

class TestExtractPipelineAttributes:
    def test_single_estimator(self):
        from risk_ml.estimator import RiskXGBClassifier
        np.random.seed(42)
        X = pd.DataFrame({"a": np.random.randn(30), "b": np.random.randn(30)})
        y = np.random.randint(0, 2, 30)
        clf = RiskXGBClassifier()
        clf.fit(X, y)
        attrs = extract_pipeline_attributes(clf)
        assert attrs.model_params_ is not None
        assert attrs.feature_names_in_ is not None

    def test_pipeline(self):
        ctx = _make_context()
        attrs = ctx.pipeline_attrs
        assert attrs is not None
        assert attrs.model_params_ is not None
