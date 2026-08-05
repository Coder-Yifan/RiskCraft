"""全流程 Demo: Pipeline 训练 + 报告产出

演示流程:
1. 加载 demo_data.csv，解析时间列
2. 最新3个月 → OOT，其余 7:3 随机 split → train/test，写入 tag 列
3. 随机生成 mob3/mob6 标签（模拟不同表现期）
4. 实验模块: 不同时间窗口训练 score1（基线）和 score2（新模型）
5. score1 作为 baseline_score，产出 score2 的标准模型报告 Excel

运行方式:
    python risk_report/demo_full_pipeline.py
"""

import sys
import io
import os
import time

# 确保项目根目录在 sys.path 中
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

# Windows GBK 终端兼容
if sys.platform == "win32":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

import numpy as np
import pandas as pd

from risk_ml import FeatureCleaner, RiskPipeline
from risk_ml.encoding import BinnerWoeEncoder
from risk_ml.feature_selection import IVSelector, CorrelationSelector
from risk_ml.estimator import RiskXGBClassifier
from risk_ml.experiment import (
    ExperimentRunner,
    ExperimentConfig,
    TimeWindow,
    AUCMetric,
    KSMetric,
    LiftMetric,
)

# ============================================================
# 1. 加载与划分数据
# ============================================================
print("=" * 70)
print("[1] 加载与划分数据")
print("=" * 70)

df = pd.read_csv("risk_ml/dataset/demo_data.csv")
df["transaction_time"] = pd.to_datetime(df["transaction_time"])

print(f"原始数据: {len(df):,} 行, 欺诈率: {df['is_fraud'].mean():.4f}")
print(f"时间范围: {df['transaction_time'].min().date()} ~ {df['transaction_time'].max().date()}")

# 最新3个月 → OOT
oot_cutoff = df["transaction_time"].max() - pd.DateOffset(months=3)
print(f"OOT 切分点: {oot_cutoff.date()}（最新3个月）")

# 标记 tag
df["tag"] = "train"  # 默认 train
df.loc[df["transaction_time"] >= oot_cutoff, "tag"] = "oot"

# 非 OOT 部分随机 7:3 split → train / test
np.random.seed(42)
non_oot_mask = df["tag"] == "train"
non_oot_idx = df.index[non_oot_mask]
test_idx = np.random.choice(non_oot_idx, size=int(len(non_oot_idx) * 0.3), replace=False)
df.loc[test_idx, "tag"] = "test"

print(f"\n数据划分:")
for tag_val in ["train", "test", "oot"]:
    mask = df["tag"] == tag_val
    n = mask.sum()
    rate = df.loc[mask, "is_fraud"].mean()
    print(f"  {tag_val:5s}: {n:>7,} 行, 欺诈率 {rate:.4f}")

# 采样 10% 加速演示
df_sample = pd.concat([
    g.sample(frac=0.1, random_state=42)
    for _, g in df.groupby("tag")
], ignore_index=True)

print(f"\n采样后（10%）:")
for tag_val in ["train", "test", "oot"]:
    mask = df_sample["tag"] == tag_val
    n = mask.sum()
    rate = df_sample.loc[mask, "is_fraud"].mean()
    print(f"  {tag_val:5s}: {n:>7,} 行, 欺诈率 {rate:.4f}")

# ============================================================
# 2. 生成 MOB 标签
# ============================================================
print("\n" + "=" * 70)
print("[2] 生成 MOB 标签（模拟不同表现期）")
print("=" * 70)

np.random.seed(42)

# y_mob3: 70% 概率与 is_fraud 一致 + 30% 随机翻转
flip_prob_mob3 = 0.30
flip_mask_mob3 = np.random.rand(len(df_sample)) < flip_prob_mob3
random_labels_mob3 = np.random.randint(0, 2, size=len(df_sample))
df_sample["y_mob3"] = np.where(
    flip_mask_mob3,
    random_labels_mob3,
    df_sample["is_fraud"],
)

# y_mob6: 60% 概率与 is_fraud 一致 + 40% 随机翻转
flip_prob_mob6 = 0.40
flip_mask_mob6 = np.random.rand(len(df_sample)) < flip_prob_mob6
random_labels_mob6 = np.random.randint(0, 2, size=len(df_sample))
df_sample["y_mob6"] = np.where(
    flip_mask_mob6,
    random_labels_mob6,
    df_sample["is_fraud"],
)

print(f"y_mob3: 坏率 {df_sample['y_mob3'].mean():.4f} (模拟3个月表现期, 翻转率{flip_prob_mob3:.0%})")
print(f"y_mob6: 坏率 {df_sample['y_mob6'].mean():.4f} (模拟6个月表现期, 翻转率{flip_prob_mob6:.0%})")
print(f"is_fraud: 坏率 {df_sample['is_fraud'].mean():.4f} (原始标签)")

# ============================================================
# 3. 构造特征列
# ============================================================
print("\n" + "=" * 70)
print("[3] 构造特征列")
print("=" * 70)

# 排除 ID/时间/标签/MOB列
exclude_cols = {
    "transaction_id", "customer_id", "transaction_time",
    "is_fraud", "y_mob3", "y_mob6", "tag",
}
feature_cols = [c for c in df_sample.columns if c not in exclude_cols]
print(f"特征数: {len(feature_cols)}")
print(f"特征列表: {feature_cols}")

# ============================================================
# 4. 实验模块: 不同时间窗口训练 score1 和 score2
# ============================================================
print("\n" + "=" * 70)
print("[4] 实验模块: 不同时间窗口训练 score1（基线）和 score2（新模型）")
print("=" * 70)

# OOT 数据（用于 ExperimentRunner 的 oot 参数）
df_oot = df_sample[df_sample["tag"] == "oot"].copy()

# 时间窗口
tw_2024 = TimeWindow("transaction_time", "2023-06-01", "2024-12-31")  # score1: 2024前
tw_2025 = TimeWindow("transaction_time", "2023-06-01", "2025-06-30")  # score2: 到2025中

configs = [
    ExperimentConfig(name="score1_baseline", label_col="is_fraud", time_window=tw_2024),
    ExperimentConfig(name="score2_new", label_col="is_fraud", time_window=tw_2025),
]

# Pipeline 不嵌套 OptunaTuner，由 ExperimentRunner 自动包装调参
pipe = RiskPipeline([
    ("cleaner", FeatureCleaner()),
    ("binner_woe", BinnerWoeEncoder(max_bins=8)),
    ("iv_selector", IVSelector(iv_threshold=0.02)),
    ("corr_selector", CorrelationSelector(corr_threshold=0.7)),
    ("classifier", RiskXGBClassifier(n_estimators=50)),
])

metrics = [AUCMetric(), KSMetric(), LiftMetric(percentile=10)]

runner = ExperimentRunner(
    configs=configs,
    pipeline=pipe,
    feature_columns=feature_cols,
    metrics=metrics,
    scoring="ks",
    n_trials=10,
    tuner_cv=3,
    oot=df_oot,
    eval_label_cols=["y_mob3", "y_mob6"],
    n_jobs=1,
    random_state=42,
    verbose=1,
)

start = time.time()
runner.fit(df_sample)
elapsed = time.time() - start
print(f"\n训练完成, 耗时 {elapsed:.1f}s")

# 展示实验结果
print("\n实验结果对比:")
display_cols = ["name", "label_col", "time_window",
                "auc", "ks", "lift_10",
                "oot_auc", "oot_ks", "oot_lift_10"]
avail_cols = [c for c in display_cols if c in runner.results_.columns]
print(runner.results_[avail_cols].to_string(index=False))

# ============================================================
# 5. 对整个 DataFrame 计算 score1 和 score2
# ============================================================
print("\n" + "=" * 70)
print("[5] 对整个 DataFrame 计算 score1 和 score2")
print("=" * 70)

# 从 experiments_ 获取两个模型的 fitted pipeline
est_score1 = runner.experiments_["score1_baseline"].estimator
est_score2 = runner.experiments_["score2_new"].estimator

# 计算预测分数
df_sample["score1_baseline"] = est_score1.predict_proba(df_sample[feature_cols])[:, 1]
df_sample["score2_new"] = est_score2.predict_proba(df_sample[feature_cols])[:, 1]

print(f"score1_baseline: mean={df_sample['score1_baseline'].mean():.4f}, "
      f"std={df_sample['score1_baseline'].std():.4f}")
print(f"score2_new:      mean={df_sample['score2_new'].mean():.4f}, "
      f"std={df_sample['score2_new'].std():.4f}")

# 各数据集的指标对比
print("\nscore1 vs score2 各数据集指标:")
from risk_report import ModelEffectOperator

for tag_val, cn_name in [("train", "训练集"), ("test", "测试集"), ("oot", "跨时间验证集")]:
    mask = df_sample["tag"] == tag_val
    y_true = df_sample.loc[mask, "is_fraud"].values
    s1 = df_sample.loc[mask, "score1_baseline"].values
    s2 = df_sample.loc[mask, "score2_new"].values

    datasets = {
        f"score1_{cn_name}": (y_true, s1),
        f"score2_{cn_name}": (y_true, s2),
    }
    df_eff = ModelEffectOperator.compute_effect_table(datasets, metrics)
    print(f"\n--- {cn_name} ---")
    print(df_eff.to_string(index=False))

# ============================================================
# 6. 构造 ReportContext 并产出报告
# ============================================================
print("\n" + "=" * 70)
print("[6] 构造 ReportContext 并产出标准模型报告")
print("=" * 70)

from risk_report import ReportContext, ModelReport

# 特征元信息
feature_meta = {
    "transaction_amount": {"含义": "交易金额", "来源": "交易系统", "类别": "数值"},
    "merchant_category": {"含义": "商户类别", "来源": "商户信息", "类别": "分类"},
    "transaction_type": {"含义": "交易类型", "来源": "交易系统", "类别": "分类"},
    "payment_method": {"含义": "支付方式", "来源": "交易系统", "类别": "分类"},
    "city": {"含义": "城市", "来源": "用户信息", "类别": "分类"},
    "country": {"含义": "国家", "来源": "用户信息", "类别": "分类"},
    "device_type": {"含义": "设备类型", "来源": "交易日志", "类别": "分类"},
    "operating_system": {"含义": "操作系统", "来源": "交易日志", "类别": "分类"},
    "browser": {"含义": "浏览器", "来源": "交易日志", "类别": "分类"},
    "card_type": {"含义": "卡类型", "来源": "卡片信息", "类别": "分类"},
    "card_present": {"含义": "是否现场刷卡", "来源": "交易系统", "类别": "二元"},
    "international_transaction": {"含义": "是否跨境交易", "来源": "交易系统", "类别": "二元"},
    "distance_from_home": {"含义": "离家距离", "来源": "位置数据", "类别": "数值"},
    "previous_transaction_gap": {"含义": "上次交易间隔", "来源": "交易日志", "类别": "数值"},
    "daily_transaction_count": {"含义": "当日交易次数", "来源": "交易日志", "类别": "数值"},
    "risk_score": {"含义": "风控评分", "来源": "风控引擎", "类别": "数值"},
    "customer_age": {"含义": "客户年龄", "来源": "客户信息", "类别": "数值"},
    "account_tenure_years": {"含义": "账户年限", "来源": "客户信息", "类别": "数值"},
    "merchant_risk_level": {"含义": "商户风险等级", "来源": "风控引擎", "类别": "分类"},
}

# 用 score2_new 作为新模型分数（pipeline 自动计算）
# 用 score1_baseline 作为对标模型分数（baseline_score_col）
context = ReportContext(
    data=df_sample,
    tag_col="tag",
    label_col="is_fraud",
    pipeline=est_score2,  # 新模型 pipeline → 自动提取属性 + 计算分数
    baseline_score_col="score1_baseline",  # 对标模型分数列
    extra_labels=["y_mob3", "y_mob6"],  # 多标签（MOB）
    time_col="transaction_time",  # 时间列（用于月度拆分分析）
    model_name="反欺诈模型_v2.0",
    developer="RiskCraft Demo",
    validator="风控团队",
    business_owner="业务运营部",
    background="针对线上交易欺诈风险，使用更多历史数据训练，构建反欺诈预测模型",
    application="线上交易实时风控筛查，辅助人工审核决策",
    label_definition={0: "正常", 1: "欺诈"},
    # observation_period="6个月",
    feature_meta=feature_meta,
)

print(f"ReportContext 构造完成:")
print(f"  score_col: {context.score_col}")
print(f"  baseline_score_col: {context.baseline_score_col}")
print(f"  extra_labels: {context.extra_labels}")
print(f"  get_datasets(): {list(context.get_datasets().keys())}")
print(f"  get_baseline_datasets(): {list(context.get_baseline_datasets().keys())}")

# 产出报告
report = ModelReport()
report.fit(context)

print(f"\n算子结果: {list(report.results_.keys())}")

output_path = "risk_report/demo_report_beautified.xlsx"
report.to_excel(output_path)
print(f"✓ Excel 已保存: {output_path}")

# 验证 Excel 内容
import openpyxl
wb = openpyxl.load_workbook(output_path)
print(f"\nExcel Sheet 概览:")
print(f"  {'Sheet名':<20s} {'行数':>6s} {'列数':>6s}")
print(f"  {'-'*20} {'-'*6} {'-'*6}")
for name in wb.sheetnames:
    ws = wb[name]
    print(f"  {name:<20s} {ws.max_row:>6d} {ws.max_column:>6d}")

# ============================================================
# 7. 验证关键算子
# ============================================================
print("\n" + "=" * 70)
print("[7] 验证关键算子")
print("=" * 70)

# score_lift: 应有 baseline 对标列
lift_result = report.get_result("score_lift")
print(f"\nscore_lift 算子: {len(lift_result)} 个 SubSection")
for sub in lift_result:
    has_baseline = any("baseline_" in c for c in sub.data.columns)
    print(f"  {sub.title}: {sub.data.shape}, 有baseline列: {has_baseline}")

# model_comparison: 新模型 vs 对标模型
comp_result = report.get_result("model_comparison")
print(f"\nmodel_comparison 算子: {len(comp_result)} 个 SubSection")
for sub in comp_result:
    print(f"  {sub.title}: {sub.data.shape}")
    if "数据集" in sub.data.columns:
        print(f"    数据集: {sub.data['数据集'].tolist()}")

# mob_performance: 多标签对比
mob_result = report.get_result("mob_performance")
print(f"\nmob_performance 算子: {len(mob_result)} 个 SubSection")
for sub in mob_result:
    print(f"  {sub.title}: {sub.data.shape}")
    if "标签列" in sub.data.columns:
        print(f"    标签列: {sub.data['标签列'].unique().tolist()}")

# model_effect: 指标对比
effect_result = report.get_result("model_effect")
print(f"\nmodel_effect 算子: {len(effect_result)} 个 SubSection")
for sub in effect_result:
    print(f"  {sub.title}: {sub.data.shape}")
    if "数据集" in sub.data.columns:
        print(f"    数据集: {sub.data['数据集'].tolist()}")

print("\n" + "=" * 70)
print("演示完成! 报告文件: risk_report/demo_full_pipeline_report.xlsx")
print("=" * 70)
