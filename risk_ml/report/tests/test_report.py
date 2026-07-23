"""risk_ml.report 模块单元测试。"""

import numpy as np
import pandas as pd
import pytest
import tempfile
import os

from risk_ml.report import (
    ReportContext, PipelineAttributes, ModelReport, ExcelWriter,
    ScoreLiftOperator, FeatureFilterSummaryOperator,
    ModelEffectOperator, SwapInOutOperator,
    compute_lift_table, compute_swap_analysis, compute_per_feature_ks, compute_sample_stats,
    MetaInfoOperator, ModelDesignOperator, VariableAnalysisOperator,
    ModelPerformanceOperator, SupplementaryOperator, UsagePlanOperator,
    VariableDescriptionOperator, SubSection, ReportSectionResult,
    FormatConfig, DEFAULT_FORMAT, extract_pipeline_attributes,
)


# ============================================================
# _scoring.py 测试
# ============================================================

class TestComputeLiftTable:
    """compute_lift_table 测试。"""

    def test_basic_output(self):
        y_true = np.array([0, 0, 0, 1, 1, 1])
        y_score = np.array([0.1, 0.3, 0.5, 0.7, 0.8, 0.9])
        df = compute_lift_table(y_true, y_score, n_bins=3)
        assert isinstance(df, pd.DataFrame)
        expected_cols = ["min", "max", "goods", "bads", "total", "total%", "bad_rate", "ks", "lift", "cum_lift"]
        for col in expected_cols:
            assert col in df.columns

    def test_with_baseline(self):
        y_true = np.array([0, 0, 1, 1])
        y_score = np.array([0.2, 0.4, 0.6, 0.8])
        baseline = np.array([0.3, 0.5, 0.5, 0.7])
        df = compute_lift_table(y_true, y_score, n_bins=2, baseline_score=baseline)
        assert "baseline_goods" in df.columns
        assert "baseline_bads" in df.columns


class TestComputeSwapAnalysis:
    """compute_swap_analysis 测试。"""

    def test_basic_no_baseline(self):
        y_true = np.array([0, 0, 1, 1])
        y_score = np.array([0.2, 0.5, 0.7, 0.9])
        df = compute_swap_analysis(y_true, y_score, cutoff_percentiles=[10])
        assert "切分比例" in df.columns
        assert "新总拒绝" in df.columns

    def test_with_baseline(self):
        y_true = np.array([0, 0, 1, 1])
        y_score_new = np.array([0.2, 0.5, 0.7, 0.9])
        y_score_old = np.array([0.3, 0.4, 0.6, 0.8])
        df = compute_swap_analysis(y_true, y_score_new, y_score_old, cutoff_percentiles=[10, 20])
        assert "swap_in" in df.columns
        assert len(df) == 2


class TestComputeSampleStats:
    """compute_sample_stats 测试。"""

    def test_binary(self):
        y = np.array([0, 0, 1, 1, 1])
        stats = compute_sample_stats(y)
        assert stats["goods"] == 2
        assert stats["bads"] == 3
        assert stats["total"] == 5
        assert stats["bad_rate"] == 0.6

    def test_with_gray(self):
        y = np.array([0, 0, -1, 1, 1])
        stats = compute_sample_stats(y, {0: "好", -1: "灰", 1: "坏"})
        assert stats["gray"] == 1
        assert stats["total"] == 4  # 不含灰


# ============================================================
# _base.py / _context.py 测试
# ============================================================

class TestReportContext:
    """ReportContext 测试。"""

    def test_empty_context(self):
        ctx = ReportContext()
        assert ctx.pipeline_attrs is None
        assert ctx.metrics is not None
        assert len(ctx.metrics) > 0

    def test_auto_score_computation(self):
        """当提供 pipeline + X_train 时，自动计算 y_score_train。"""
        np.random.seed(42)
        n = 300
        X = pd.DataFrame({
            "f1": np.random.randn(n),
            "f2": np.random.randn(n),
            "f3": np.random.randn(n),
            "f4": np.random.randn(n),
            "f5": np.random.randn(n),
        })
        # 让 y 与特征有相关性，确保 IV 筛选后保留特征
        y = (X["f1"] + X["f2"] > 0).astype(int)

        from sklearn.pipeline import Pipeline
        from risk_ml.preprocessing import FeatureCleaner
        from risk_ml.encoding import BinnerWoeEncoder
        from risk_ml.feature_selection import IVSelector
        from risk_ml.estimator import RiskXGBClassifier

        pipe = Pipeline([
            ("cleaner", FeatureCleaner()),
            ("binner_woe", BinnerWoeEncoder()),
            ("iv_selector", IVSelector()),
            ("classifier", RiskXGBClassifier()),
        ])
        pipe.fit(X, y)

        ctx = ReportContext(pipeline=pipe, X_train=X, y_train=y)
        assert ctx.y_score_train is not None
        assert ctx.pipeline_attrs is not None


class TestExtractPipelineAttributes:
    """extract_pipeline_attributes 测试。"""

    def test_single_estimator(self):
        from risk_ml.estimator import RiskXGBClassifier
        np.random.seed(42)
        X = pd.DataFrame({"a": np.random.randn(30), "b": np.random.randn(30)})
        y = np.random.randint(0, 2, 30)
        clf = RiskXGBClassifier()
        clf.fit(X, y)
        attrs = extract_pipeline_attributes(clf)
        assert attrs.model_params_ is not None
        assert attrs.feature_names_in_ is not None


# ============================================================
# ExcelWriter 测试
# ============================================================

class TestExcelWriter:
    """ExcelWriter 测试。"""

    def test_write_dataframe(self):
        df = pd.DataFrame({"col_a": [1, 2, 3], "col_b": [0.1, 0.2, 0.3]})
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            writer = ExcelWriter(path)
            writer.write_dataframe(df, "test_sheet", title="测试标题")
            writer.save()

            import openpyxl
            wb = openpyxl.load_workbook(path)
            assert "test_sheet" in wb.sheetnames
            ws = wb["test_sheet"]
            assert ws.cell(row=1, column=1).value == "测试标题"
            assert ws.cell(row=2, column=1).value == "col_a"
        finally:
            os.unlink(path)

    def test_write_section(self):
        section = ReportSectionResult(
            sheet_name="测试",
            sub_sections=[
                SubSection(title="子章节1", data=pd.DataFrame({"x": [1, 2]})),
                SubSection(title="子章节2", data=pd.DataFrame({"y": [3, 4]})),
            ],
        )
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            writer = ExcelWriter(path)
            writer.write_section(section)
            writer.save()

            import openpyxl
            wb = openpyxl.load_workbook(path)
            assert "测试" in wb.sheetnames
        finally:
            os.unlink(path)


# ============================================================
# 算子测试
# ============================================================

class TestScoreLiftOperator:
    """ScoreLiftOperator 测试。"""

    def test_static_method(self):
        y_true = np.array([0, 0, 1, 1])
        y_score = np.array([0.2, 0.4, 0.6, 0.8])
        df = ScoreLiftOperator.compute_lift_table(y_true, y_score, n_bins=2)
        assert isinstance(df, pd.DataFrame)
        assert "min" in df.columns

    def test_compute_via_context(self):
        np.random.seed(42)
        n = 300
        X = pd.DataFrame({"f1": np.random.randn(n), "f2": np.random.randn(n), "f3": np.random.randn(n), "f4": np.random.randn(n), "f5": np.random.randn(n)})
        y = (X["f1"] + X["f2"] > 0).astype(int)
        from sklearn.pipeline import Pipeline
        from risk_ml.preprocessing import FeatureCleaner
        from risk_ml.encoding import BinnerWoeEncoder
        from risk_ml.feature_selection import IVSelector
        from risk_ml.estimator import RiskXGBClassifier
        pipe = Pipeline([("cleaner", FeatureCleaner()), ("binner_woe", BinnerWoeEncoder()), ("iv_selector", IVSelector()), ("classifier", RiskXGBClassifier())])
        pipe.fit(X, y)
        ctx = ReportContext(pipeline=pipe, X_train=X, y_train=y)
        op = ScoreLiftOperator(n_bins=5)
        result = op.compute(ctx)
        assert result.sheet_name == "模型分分箱表现"
        assert len(result.sub_sections) > 0


class TestModelEffectOperator:
    """ModelEffectOperator 测试。"""

    def test_static_method(self):
        datasets = {"train": (np.array([0, 0, 1, 1]), np.array([0.2, 0.4, 0.6, 0.8]))}
        df = ModelEffectOperator.compute_effect_table(datasets)
        assert "数据集" in df.columns
        assert "auc" in df.columns


class TestSwapInOutOperator:
    """SwapInOutOperator 测试。"""

    def test_static_method(self):
        y_true = np.array([0, 0, 1, 1])
        y_score = np.array([0.2, 0.5, 0.7, 0.9])
        df = SwapInOutOperator.compute_swap_table(y_true, y_score)
        assert "切分比例" in df.columns


class TestMetaInfoOperator:
    """MetaInfoOperator 测试。"""

    def test_compute(self):
        ctx = ReportContext(model_name="测试", developer="dev")
        op = MetaInfoOperator()
        result = op.compute(ctx)
        assert result.sheet_name == "模型说明"
        assert len(result.sub_sections) == 1
        df = result.sub_sections[0].data
        assert "项目" in df.columns


class TestModelDesignOperator:
    """ModelDesignOperator 测试。"""

    def test_compute_minimal(self):
        ctx = ReportContext(model_name="模型A", background="降低违约率")
        op = ModelDesignOperator()
        result = op.compute(ctx)
        assert result.sheet_name == "1.模型设计"
        assert len(result.sub_sections) >= 3  # 至少有 1.1/1.2/1.3


# ============================================================
# ModelReport 组合器测试
# ============================================================

class TestModelReport:
    """ModelReport 组合器测试。"""

    def test_default_operators(self):
        report = ModelReport()
        assert len(report.operators) == 7

    def test_custom_operators(self):
        report = ModelReport(operators=[MetaInfoOperator(), ScoreLiftOperator()])
        assert len(report.operators) == 2

    def test_fit_and_to_excel(self):
        """完整流程: fit → to_excel。"""
        np.random.seed(42)
        n = 300
        X = pd.DataFrame({"f1": np.random.randn(n), "f2": np.random.randn(n), "f3": np.random.randn(n), "f4": np.random.randn(n), "f5": np.random.randn(n)})
        y = (X["f1"] + X["f2"] > 0).astype(int)
        from sklearn.pipeline import Pipeline
        from risk_ml.preprocessing import FeatureCleaner
        from risk_ml.encoding import BinnerWoeEncoder
        from risk_ml.feature_selection import IVSelector
        from risk_ml.estimator import RiskXGBClassifier
        pipe = Pipeline([("cleaner", FeatureCleaner()), ("binner_woe", BinnerWoeEncoder()), ("iv_selector", IVSelector()), ("classifier", RiskXGBClassifier())])
        pipe.fit(X, y)

        ctx = ReportContext(pipeline=pipe, X_train=X, y_train=y)
        report = ModelReport(operators=[MetaInfoOperator(), ModelDesignOperator()])
        report.fit(ctx)

        assert "meta_info" in report.results_
        assert "model_design" in report.results_

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            report.to_excel(path)
            import openpyxl
            wb = openpyxl.load_workbook(path)
            assert len(wb.sheetnames) == 2
        finally:
            os.unlink(path)

    def test_get_section(self):
        ctx = ReportContext(model_name="test")
        report = ModelReport(operators=[MetaInfoOperator()])
        report.fit(ctx)
        section = report.get_section("meta_info")
        assert section.sheet_name == "模型说明"

    def test_get_section_not_found(self):
        ctx = ReportContext(model_name="test")
        report = ModelReport(operators=[MetaInfoOperator()])
        report.fit(ctx)
        with pytest.raises(KeyError):
            report.get_section("nonexistent")
